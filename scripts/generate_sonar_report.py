#!/usr/bin/env python3
import os
import sys
import io
import json
import datetime
import base64
import urllib.request
import urllib.error
import urllib.parse
from pathlib import Path

# Fix encoding cho Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from dotenv import load_dotenv

# Load env variables from .env if local
load_dotenv()

# --- CONFIGURATION ---
PROJECTS = {
    "Backend (Spring Boot)": "novaticket-backend",
    "Frontend (React)": "novaticket-frontend",
    "Android App": "novaticket-android",
    "AI Service (FastAPI)": "novaticket-ai"
}

# --- RATING MAPPING ---
def map_rating(value):
    if value is None:
        return "N/A"
    val_str = str(value).strip()
    try:
        val_float = float(val_str)
        val_int = int(val_float)
        mapping = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E"}
        return mapping.get(val_int, "N/A")
    except ValueError:
        return "N/A"

# --- HTTP REQUEST HELPER ---
def send_request(url, headers=None):
    if headers is None:
        headers = {}
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req) as response:
            res_data = response.read().decode("utf-8")
            return json.loads(res_data)
    except urllib.error.HTTPError as e:
        err_content = e.read().decode("utf-8")
        print(f"❌ HTTP Error {e.code} on {url}: {e.reason}")
        print(f"   Response: {err_content}")
        return {}
    except Exception as e:
        print(f"❌ Connection Error on {url}: {e}")
        return {}

# --- FETCH SONAR HEALTH DATA ---
def fetch_project_health(project_key, sonar_token):
    auth_bytes = f"{sonar_token}:".encode("utf-8")
    sonar_auth = f"Basic {base64.b64encode(auth_bytes).decode('utf-8')}"
    headers = {"Authorization": sonar_auth, "Accept": "application/json"}
    
    # 1. Quality Gate Status
    qg_url = f"https://sonarcloud.io/api/qualitygates/project_status?projectKey={project_key}"
    qg_res = send_request(qg_url, headers=headers)
    status = qg_res.get("projectStatus", {}).get("status", "N/A")
    qg_status = "Passed" if status == "OK" else ("Failed" if status == "ERROR" else "N/A")
    
    # 2. Key Metrics
    metrics_url = f"https://sonarcloud.io/api/measures/component?component={project_key}&metricKeys=bugs,vulnerabilities,code_smells,coverage,duplicated_lines_density,reliability_rating,security_rating,sqale_rating"
    metrics_res = send_request(metrics_url, headers=headers)
    
    measures_list = metrics_res.get("component", {}).get("measures", [])
    metrics = {m.get("metric"): m.get("value") for m in measures_list}
    
    # Format metrics
    bugs = metrics.get("bugs", "0")
    vulnerabilities = metrics.get("vulnerabilities", "0")
    code_smells = metrics.get("code_smells", "0")
    
    coverage_val = metrics.get("coverage")
    coverage = f"{float(coverage_val):.1f}%" if coverage_val is not None else "N/A"
    
    dup_val = metrics.get("duplicated_lines_density")
    duplication = f"{float(dup_val):.1f}%" if dup_val is not None else "N/A"
    
    reliability = map_rating(metrics.get("reliability_rating"))
    security = map_rating(metrics.get("security_rating"))
    maintainability = map_rating(metrics.get("sqale_rating"))
    
    return {
        "qg_status": qg_status,
        "bugs": int(bugs) if bugs.isdigit() else 0,
        "vulnerabilities": int(vulnerabilities) if vulnerabilities.isdigit() else 0,
        "code_smells": int(code_smells) if code_smells.isdigit() else 0,
        "coverage": coverage,
        "duplication": duplication,
        "reliability": reliability,
        "security": security,
        "maintainability": maintainability
    }

# --- FETCH WEEKLY RESOLVED ISSUES ---
def fetch_weekly_fixed_issues(project_key, sonar_token):
    auth_bytes = f"{sonar_token}:".encode("utf-8")
    sonar_auth = f"Basic {base64.b64encode(auth_bytes).decode('utf-8')}"
    headers = {"Authorization": sonar_auth, "Accept": "application/json"}
    
    # Fetch resolved issues
    url = f"https://sonarcloud.io/api/issues/search?componentKeys={project_key}&resolved=true&resolutions=FIXED&types=BUG,VULNERABILITY,CODE_SMELL&ps=500"
    res = send_request(url, headers=headers)
    issues_list = res.get("issues", [])
    
    weekly_fixed = []
    now_time = datetime.datetime.utcnow()
    
    for issue in issues_list:
        update_date_str = issue.get("updateDate", "")
        if not update_date_str:
            continue
        try:
            naive_str = update_date_str[:19]
            dt = datetime.datetime.strptime(naive_str, "%Y-%m-%dT%H:%M:%S")
            delta = now_time - dt
            # Filter issues fixed within last 7 days
            if delta.days <= 7:
                comp = issue.get("component", "")
                parts = comp.split(":", 1)
                file_path = parts[1] if len(parts) > 1 else comp
                
                weekly_fixed.append({
                    "file_path": file_path,
                    "type": issue.get("type", "CODE_SMELL").replace("_", " ").title(),
                    "message": issue.get("message", ""),
                    "severity": issue.get("severity", "MAJOR").title(),
                    "fixed_date": dt.strftime("%d/%m/%Y")
                })
        except Exception as e:
            print(f"⚠️ Error filtering date {update_date_str} for issue {issue.get('key')}: {e}")
            
    return weekly_fixed

# --- BUILD EXCEL WORKBOOK ---
def build_report(sonar_token):
    wb = openpyxl.Workbook()
    
    # Fonts and fills
    font_title = Font(name="Inter", size=16, bold=True, color="0D1B2A")
    font_header = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Inter", size=10)
    font_bold = Font(name="Inter", size=10, bold=True)
    
    font_passed = Font(name="Inter", size=10, bold=True, color="2E7D32")
    font_failed = Font(name="Inter", size=10, bold=True, color="C62828")
    
    fill_header = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    fill_passed = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_failed = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    fill_a = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_e = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    font_a = Font(name="Inter", size=10, bold=True, color="2E7D32")
    font_e = Font(name="Inter", size=10, bold=True, color="C62828")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # --- SHEET 1: PROJECT HEALTH ---
    ws1 = wb.active
    ws1.title = "Project Health"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"BÁO CÁO SỨC KHỎE MÃ NGUỒN SONARCLOUD - {datetime.date.today().strftime('%d/%m/%Y')}"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40
    
    headers1 = [
        "Component / Phân Hệ", "Quality Gate", "Bugs", 
        "Vulnerabilities", "Code Smells", "Test Coverage", 
        "Duplication", "Reliability Rating", "Security Rating", "Maintainability Rating"
    ]
    ws1.row_dimensions[3].height = 28
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    row_idx = 4
    all_resolved = []
    
    for name, key in PROJECTS.items():
        print(f"📊 Đang truy xuất dữ liệu Sonar cho: {name} ({key})...")
        health = fetch_project_health(key, sonar_token)
        
        ws1.row_dimensions[row_idx].height = 24
        ws1.cell(row=row_idx, column=1, value=name).font = font_bold
        ws1.cell(row=row_idx, column=1).alignment = Alignment(vertical="center", indent=1)
        
        # Quality Gate
        cell_qg = ws1.cell(row=row_idx, column=2, value=health["qg_status"])
        cell_qg.alignment = Alignment(horizontal="center", vertical="center")
        if health["qg_status"] == "Passed":
            cell_qg.fill = fill_passed
            cell_qg.font = font_passed
        elif health["qg_status"] == "Failed":
            cell_qg.fill = fill_failed
            cell_qg.font = font_failed
        else:
            cell_qg.font = font_body
            
        # Standard values
        ws1.cell(row=row_idx, column=3, value=health["bugs"])
        ws1.cell(row=row_idx, column=4, value=health["vulnerabilities"])
        ws1.cell(row=row_idx, column=5, value=health["code_smells"])
        ws1.cell(row=row_idx, column=6, value=health["coverage"])
        ws1.cell(row=row_idx, column=7, value=health["duplication"])
        
        # Ratings
        for col_idx, rating_key in enumerate(["reliability", "security", "maintainability"], 8):
            cell_rt = ws1.cell(row=row_idx, column=col_idx, value=health[rating_key])
            cell_rt.alignment = Alignment(horizontal="center", vertical="center")
            if health[rating_key] == "A":
                cell_rt.fill = fill_a
                cell_rt.font = font_a
            elif health[rating_key] in ["D", "E"]:
                cell_rt.fill = fill_e
                cell_rt.font = font_e
            else:
                cell_rt.font = font_bold
                
        # Format alignment and border
        for c in range(1, 11):
            cell = ws1.cell(row=row_idx, column=c)
            cell.border = thin_border
            if c > 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        # Accumulate resolved issues
        resolved = fetch_weekly_fixed_issues(key, sonar_token)
        for issue in resolved:
            issue["project"] = name
        all_resolved.extend(resolved)
        
        row_idx += 1
        
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 16
    for c in ["C", "D", "E", "F", "G", "H", "I", "J"]:
        ws1.column_dimensions[c].width = 15
        
    # --- SHEET 2: WEEKLY FIXED BUGS ---
    ws2 = wb.create_sheet(title="Weekly Fixed Bugs")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"DANH SÁCH LỖI ĐÃ ĐƯỢC FIX TRONG TUẦN - {datetime.date.today().strftime('%d/%m/%Y')}"
    ws2["A1"].font = font_title
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40
    
    headers2 = ["Phân Hệ", "Đường Dẫn File", "Loại Lỗi", "Mô Tả Lỗi", "Mức Độ Cảnh Báo", "Ngày Sửa Xong"]
    ws2.row_dimensions[3].height = 28
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    row_idx = 4
    if not all_resolved:
        ws2.merge_cells("A4:F4")
        ws2["A4"] = "🎉 Tuyệt vời! Tuần này không có lỗi phát sinh mới cần sửa hoặc toàn bộ đã sạch lỗi."
        ws2["A4"].font = font_bold
        ws2["A4"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[4].height = 36
        ws2.cell(row=4, column=1).border = thin_border
        for c in range(1, 7):
            ws2.cell(row=4, column=c).border = thin_border
    else:
        for issue in all_resolved:
            ws2.row_dimensions[row_idx].height = 40
            ws2.cell(row=row_idx, column=1, value=issue["project"]).font = font_bold
            ws2.cell(row=row_idx, column=2, value=issue["file_path"])
            ws2.cell(row=row_idx, column=3, value=issue["type"])
            ws2.cell(row=row_idx, column=4, value=issue["message"])
            ws2.cell(row=row_idx, column=5, value=issue["severity"])
            ws2.cell(row=row_idx, column=6, value=issue["fixed_date"])
            
            for c in range(1, 7):
                cell = ws2.cell(row=row_idx, column=c)
                cell.font = font_body
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if c in [1, 3, 5, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1
            
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 45
    ws2.column_dimensions["E"].width = 16
    ws2.column_dimensions["F"].width = 16
    
    # Save Local Report
    out_dir = Path("baocaoLocal")
    out_dir.mkdir(exist_ok=True)
    report_name = f"NovaTicket_Sonar_Report_Week_{datetime.date.today().strftime('%U')}.xlsx"
    file_path = out_dir / report_name
    wb.save(file_path)
    print(f"🎉 Đã xuất thành công file báo cáo Excel tại: {file_path}")
    return str(file_path)

# --- UPLOAD TO GDRIVE (REUSE VERIFIED LOGIC) ---
def upload_to_gdrive(file_path):
    folder_id = os.getenv("GDRIVE_FOLDER_ID")
    if not folder_id:
        print("⚠️ Thiếu cấu hình thư mục Google Drive (GDRIVE_FOLDER_ID). Không thể upload!")
        return

    file_name = os.path.basename(file_path)
    file_metadata = {
        'name': file_name,
        'parents': [folder_id]
    }
    media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

    # 1. OAuth2 Fallback Authorization
    refresh_token = os.getenv("GDRIVE_REFRESH_TOKEN")
    client_id = os.getenv("GDRIVE_CLIENT_ID")
    client_secret = os.getenv("GDRIVE_CLIENT_SECRET")

    if refresh_token and client_id and client_secret:
        print("🔄 Đang thử xác thực và upload bằng OAuth2 (Tài khoản cá nhân)...")
        from google.oauth2.credentials import Credentials
        try:
            creds = Credentials(
                token=None,
                refresh_token=refresh_token,
                token_uri="https://oauth2.googleapis.com/token",
                client_id=client_id,
                client_secret=client_secret
            )
            service = build('drive', 'v3', credentials=creds)
            service.about().get(fields="user").execute() # Verify network
            
            new_file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            print(f"✅ Đã tải thành công file báo cáo lên Google Drive bằng OAuth2!")
            print(f" Tên file: {file_name} | ID: {new_file.get('id')}")
            return
        except Exception as e:
            print(f"⚠️ OAuth2 gặp lỗi xác thực: {e}")
            print("🔄 Đang tự động chuyển sang Service Account...")

    # 2. Service Account Authorization
    key_path = ".gdrive_key.json"
    raw_key = os.getenv("GDRIVE_SERVICE_ACCOUNT_KEY_RAW")
    if raw_key:
        with open(key_path, "w") as f:
            f.write(raw_key)

    if os.path.exists(key_path):
        try:
            print("🔄 Đang tiến hành upload bằng Service Account...")
            scopes = ['https://www.googleapis.com/auth/drive']
            creds = service_account.Credentials.from_service_account_file(key_path, scopes=scopes)
            service = build('drive', 'v3', credentials=creds)
            
            new_file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            print(f"✅ Đã tải thành công file báo cáo lên Google Drive bằng Service Account!")
            print(f" Tên file: {file_name} | ID: {new_file.get('id')}")
            
            if os.getenv("GDRIVE_SERVICE_ACCOUNT_KEY_RAW") and os.path.exists(key_path):
                try:
                    os.remove(key_path)
                except Exception:
                    pass
            return
        except Exception as sa_error:
            print(f"❌ Lỗi khi upload bằng Service Account: {sa_error}")
    else:
        print("❌ Không tìm thấy thông tin xác thực Google Drive hợp lệ!")

# --- MAIN RUN ---
def main():
    sonar_token = os.getenv("SONAR_TOKEN")
    if not sonar_token:
        print("❌ Thiếu biến môi trường SONAR_TOKEN")
        sys.exit(1)
        
    report_file = build_report(sonar_token)
    if report_file:
        upload_to_gdrive(report_file)

if __name__ == "__main__":
    main()
