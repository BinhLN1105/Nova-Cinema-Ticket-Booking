import os
import sys
import io
import json
import datetime

# Fix encoding cho Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from dotenv import load_dotenv

# Load cấu hình .env
load_dotenv()

def generate_academic_report():
    security_collection_path = "qa-tests/postman/NOVATicket_Security.postman_collection.json"
    bva_collection_path = "qa-tests/postman/NOVATicket_BVA.postman_collection.json"
    
    security_results_path = "baocaoLocal/security-report.json"
    bva_results_path = "baocaoLocal/bva-report.json"
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_dir = "baocaoLocal"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    excel_output = os.path.join(output_dir, f"NovaTicket_Academic_Test_Report_{timestamp}.xlsx")

    # 1. Đọc kết quả chạy thực tế từ Newman
    def load_actual_results(results_path):
        results = {}
        if os.path.exists(results_path):
            try:
                with open(results_path, "r", encoding="utf-8") as f:
                    report_data = json.load(f)
                    for execution in report_data.get("run", {}).get("executions", []):
                        item_name = execution.get("item", {}).get("name")
                        is_failed = False
                        assertions_log = []
                        for assertion in execution.get("assertions", []):
                            err = assertion.get("error")
                            if err:
                                is_failed = True
                                assertions_log.append(f"❌ {assertion.get('assertion')}: {err.get('message')}")
                            else:
                                assertions_log.append(f"✅ {assertion.get('assertion')}")
                        results[item_name] = {
                            "status": "Failed" if is_failed else "Passed",
                            "log": "\n".join(assertions_log)
                        }
                print(f"📊 Đã nạp thành công kết quả từ {results_path}")
            except Exception as e:
                print(f"⚠️ Không thể đọc kết quả {results_path}: {e}")
        return results

    security_actual = load_actual_results(security_results_path)
    bva_actual = load_actual_results(bva_results_path)

    # 2. Khởi tạo Workbook
    wb = openpyxl.Workbook()
    
    # Định dạng style chung
    font_title = Font(name="Inter", size=16, bold=True, color="0D1B2A")
    font_section = Font(name="Inter", size=12, bold=True, color="0D1B2A")
    font_header = Font(name="Inter", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="Inter", size=9)
    font_passed = Font(name="Inter", size=9, bold=True, color="2E7D32")
    font_failed = Font(name="Inter", size=9, bold=True, color="C62828")
    font_untested = Font(name="Inter", size=9, bold=True, color="757575")

    fill_header = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    fill_section = PatternFill(start_color="F5C518", end_color="F5C518", fill_type="solid") # Gold accent
    fill_passed = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_failed = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    fill_untested = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # ==========================================
    # SHEET 1: BVA TEST REPORT
    # ==========================================
    ws_bva = wb.active
    ws_bva.title = "BVA Test Report"
    ws_bva.views.sheetView[0].showGridLines = True

    # Title
    ws_bva.merge_cells("A1:H1")
    ws_bva["A1"] = f"BÁO CÁO KIỂM THỬ PHÂN TÍCH GIÁ TRỊ BIÊN (BVA) - NOVA TICKET"
    ws_bva["A1"].font = font_title
    ws_bva["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_bva.row_dimensions[1].height = 35

    # 1. BẢNG LỚP TƯƠNG ĐƯƠNG (Equivalent Partitioning)
    ws_bva["A3"] = "1. BẢNG PHÂN HOẠCH LỚP TƯƠNG ĐƯƠNG"
    ws_bva["A3"].font = font_section
    ws_bva.row_dimensions[3].height = 20

    headers_ep = ["Biến đầu vào / Chức năng", "Lớp tương đương Hợp lệ", "Tag", "Lớp tương đương Không hợp lệ", "Tag"]
    ws_bva.row_dimensions[4].height = 24
    for col_num, h in enumerate(headers_ep, 1):
        cell = ws_bva.cell(row=4, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ep_data = [
        ["Đăng ký: Mật khẩu (Password)", "Độ dài 6-32 ký tự, chứa cả chữ cái và chữ số", "V1", "Độ dài < 6 hoặc > 32; hoặc chỉ có chữ/chỉ có số", "X1"],
        ["Đăng ký: Họ tên (FullName)", "Độ dài từ 2 đến 100 ký tự", "V2", "Độ dài < 2 hoặc > 100 ký tự", "X2"],
        ["Đánh giá: Điểm số (Rating)", "1 <= Rating <= 5", "V3", "Rating < 1\nRating > 5", "X3\nX4"],
        ["Đặt vé: Số lượng ghế / Vé", "1 <= Ghế <= 6 (Cấu hình hệ thống)", "V4", "Ghế = 0 (Trống)\nGhế > 6 (Vượt hạn mức)", "X5\nX6"],
        ["Đặt vé: Số lượng combo / Vé", "0 <= Combo <= 8 (Cấu hình hệ thống)", "V5", "Combo < 0\nCombo > 8 (Vượt hạn mức)", "X7\nX8"]
    ]

    curr_row = 5
    for row in ep_data:
        ws_bva.row_dimensions[curr_row].height = 32
        for col_idx, val in enumerate(row, 1):
            cell = ws_bva.cell(row=curr_row, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in [3, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        curr_row += 1

    # 2. BẢNG PHÂN TÍCH GIÁ TRỊ BIÊN (BVA)
    curr_row += 1
    ws_bva.cell(row=curr_row, column=1, value="2. BẢNG PHÂN TÍCH GIÁ TRỊ BIÊN (STANDARD BVA)").font = font_section
    ws_bva.row_dimensions[curr_row].height = 20
    curr_row += 1

    headers_bva = ["Biến đầu vào", "Min (Biên dưới HL)", "Min+ (Sát biên HL)", "Nominal (Đại diện)", "Max- (Sát biên HL)", "Max (Biên trên HL)", "Biên không hợp lệ"]
    ws_bva.row_dimensions[curr_row].height = 24
    for col_num, h in enumerate(headers_bva, 1):
        cell = ws_bva.cell(row=curr_row, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    bva_data = [
        ["Độ dài Mật khẩu", "6 ký tự (HL)", "7 ký tự (HL)", "12 ký tự (HL)", "31 ký tự (HL)", "32 ký tự (HL)", "5 ký tự (KHL), 33 ký tự (KHL), Chỉ có chữ, Chỉ có số"],
        ["Độ dài Họ tên", "2 ký tự (HL)", "3 ký tự (HL)", "15 ký tự (HL)", "99 ký tự (HL)", "100 ký tự (HL)", "1 ký tự (KHL), 101 ký tự (KHL)"],
        ["Điểm số Rating", "1 (HL)", "2 (HL)", "3 (HL)", "4 (HL)", "5 (HL)", "0 (KHL), 6 (KHL)"],
        ["Hạn mức Ghế", "1 ghế (HL)", "2 ghế (HL)", "3 ghế (HL)", "5 ghế (HL)", "6 ghế (HL)", "7 ghế (KHL)"],
        ["Hạn mức Combo", "0 combo (HL)", "1 combo (HL)", "4 combo (HL)", "7 combo (HL)", "8 combo (HL)", "9 combo (KHL)"]
    ]

    curr_row += 1
    for row in bva_data:
        ws_bva.row_dimensions[curr_row].height = 28
        for col_idx, val in enumerate(row, 1):
            cell = ws_bva.cell(row=curr_row, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        curr_row += 1

    # 3. KẾT QUẢ CHẠY TEST CASE THỰC TẾ
    curr_row += 1
    ws_bva.cell(row=curr_row, column=1, value="3. BẢNG KẾT QUẢ CHẠY TEST CASE THỰC TẾ (BVA)").font = font_section
    ws_bva.row_dimensions[curr_row].height = 20
    curr_row += 1

    headers_run = ["STT", "Tên Test Case BVA", "Phương thức / Endpoint", "Request Body (Dữ liệu test)", "Kết quả mong đợi", "Kết quả thực tế (Newman)", "Trạng thái", "Tag bao phủ"]
    ws_bva.row_dimensions[curr_row].height = 28
    for col_num, h in enumerate(headers_run, 1):
        cell = ws_bva.cell(row=curr_row, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Parse collection BVA để lấy danh sách requests
    bva_testcases = []
    if os.path.exists(bva_collection_path):
        try:
            with open(bva_collection_path, "r", encoding="utf-8") as f:
                coll = json.load(f)
                idx = 1
                for folder in coll.get("item", []):
                    folder_name = folder.get("name")
                    for req_item in folder.get("item", []):
                        req = req_item.get("request", {})
                        name = req_item.get("name", "")
                        method = req.get("method", "POST")
                        path = "/".join(req.get("url", {}).get("path", []))
                        body_raw = req.get("body", {}).get("raw", "")
                        
                        # Expected
                        expected = "API phản hồi hợp lệ."
                        if "không hợp lệ" in name.lower() or "vượt biên" in name.lower() or "thiếu" in name.lower() or "bị chặn" in name.lower():
                            expected = "Bị chặn lỗi 400 Bad Request."

                        # Lấy kết quả thực tế từ newman run
                        res_info = bva_actual.get(name, {"status": "Untested", "log": "Chưa chạy kiểm thử"})
                        
                        # Tag
                        tag = "N/A"
                        if "BVA_PWD_01" in name: tag = "X1 (Password < 6)"
                        elif "BVA_PWD_02" in name: tag = "V1 (Password = 6)"
                        elif "BVA_PWD_03" in name: tag = "V1 (Password Nominal)"
                        elif "BVA_PWD_04" in name: tag = "V1 (Password = 32)"
                        elif "BVA_PWD_05" in name: tag = "X1 (Password > 32)"
                        elif "BVA_PWD_06" in name: tag = "X1 (Password thiếu số)"
                        elif "BVA_PWD_07" in name: tag = "X1 (Password thiếu chữ)"
                        elif "BVA_NAM_01" in name: tag = "X2 (Name < 2)"
                        elif "BVA_NAM_02" in name: tag = "V2 (Name = 2)"
                        elif "BVA_NAM_03" in name: tag = "V2 (Name = 100)"
                        elif "BVA_NAM_04" in name: tag = "X2 (Name > 100)"
                        elif "BVA_RAT_01" in name: tag = "X3 (Rating = 0)"
                        elif "BVA_RAT_02" in name: tag = "V3 (Rating = 1)"
                        elif "BVA_RAT_03" in name: tag = "V3 (Rating = 5)"
                        elif "BVA_RAT_04" in name: tag = "X4 (Rating = 6)"
                        elif "BVA_SEAT_01" in name: tag = "V4 (Seat = 6)"
                        elif "BVA_SEAT_02" in name: tag = "X6 (Seat = 7)"
                        elif "BVA_COMBO_01" in name: tag = "V5 (Combo = 8)"
                        elif "BVA_COMBO_02" in name: tag = "X8 (Combo = 9)"

                        bva_testcases.append({
                            "stt": idx,
                            "name": name,
                            "endpoint": f"[{method}] /{path}",
                            "body": body_raw,
                            "expected": expected,
                            "actual": res_info["log"],
                            "status": res_info["status"],
                            "tag": tag
                        })
                        idx += 1
        except Exception as e:
            print(f"⚠️ Lỗi khi trích xuất collection BVA: {e}")

    curr_row += 1
    for tc in bva_testcases:
        ws_bva.row_dimensions[curr_row].height = 65
        ws_bva.cell(row=curr_row, column=1, value=tc["stt"])
        ws_bva.cell(row=curr_row, column=2, value=tc["name"])
        ws_bva.cell(row=curr_row, column=3, value=tc["endpoint"])
        ws_bva.cell(row=curr_row, column=4, value=tc["body"])
        ws_bva.cell(row=curr_row, column=5, value=tc["expected"])
        ws_bva.cell(row=curr_row, column=6, value=tc["actual"])
        
        status_cell = ws_bva.cell(row=curr_row, column=7, value=tc["status"])
        ws_bva.cell(row=curr_row, column=8, value=tc["tag"])

        # Format styles
        for col_idx in range(1, 9):
            cell = ws_bva.cell(row=curr_row, column=col_idx)
            cell.font = font_body
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in [1, 3, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color status
        if tc["status"] == "Passed":
            status_cell.fill = fill_passed
            status_cell.font = font_passed
        elif tc["status"] == "Failed":
            status_cell.fill = fill_failed
            status_cell.font = font_failed
        else:
            status_cell.fill = fill_untested
            status_cell.font = font_untested

        curr_row += 1

    # Set column widths
    widths_bva = {"A": 8, "B": 24, "C": 22, "D": 32, "E": 20, "F": 35, "G": 12, "H": 15}
    for col, w in widths_bva.items():
        ws_bva.column_dimensions[col].width = w

    # ==========================================
    # SHEET 2: SECURITY TEST REPORT
    # ==========================================
    ws_sec = wb.create_sheet(title="Security Test Report")
    ws_sec.views.sheetView[0].showGridLines = True

    # Title
    ws_sec.merge_cells("A1:G1")
    ws_sec["A1"] = f"BÁO CÁO KIỂM THỬ BẢO MẬT (SECURITY TESTING) - NOVA TICKET"
    ws_sec["A1"].font = font_title
    ws_sec["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sec.row_dimensions[1].height = 35

    # 1. MÔ TẢ CÁC BIỆN PHÁP BẢO MẬT ĐÃ TRIỂN KHAI
    ws_sec["A3"] = "1. CÁC BIỆN PHÁP BẢO MẬT ĐÃ ĐƯỢC THIẾT LẬP"
    ws_sec["A3"].font = font_section
    ws_sec.row_dimensions[3].height = 20

    headers_sec_desc = ["Hạng mục bảo mật", "Chi tiết cơ chế hiện thực", "Mục tiêu kiểm thử"]
    ws_sec.row_dimensions[4].height = 24
    for col_num, h in enumerate(headers_sec_desc, 1):
        cell = ws_sec.cell(row=4, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    sec_desc_data = [
        ["XSS Sanitizer (Jsoup)", "Sử dụng Jackson Deserializer toàn cục kết hợp thư viện Jsoup để tự động làm sạch (clean) các thẻ HTML độc hại trong payload dạng JSON.", "Đảm bảo mã độc <script> hoặc <img> onerror không thể lọt vào database, tự động lọc sạch chuỗi."],
        ["Rate Limiting (AOP + Redis)", "Thiết lập Aspect @RateLimit với bộ đếm Redis nguyên tử để giới hạn tần suất gửi yêu cầu dựa trên IP hoặc User ID trong một khoảng thời gian.", "Đảm bảo API nhạy cảm (Đăng nhập, Thanh toán, Đặt vé) không bị tấn công Brute Force hay DDoS spam."],
        ["SQL Injection Protection", "Sử dụng Spring Data JPA (Hibernate) với cơ chế Parameterized Queries (truy vấn tham số hóa) tự động ép kiểu và escaping dữ liệu.", "Đảm bảo các ký tự đặc biệt như ' hoặc -- không làm thay đổi cấu trúc câu lệnh SQL và không rò rỉ dữ liệu DB."]
    ]

    curr_row = 5
    for row in sec_desc_data:
        ws_sec.row_dimensions[curr_row].height = 40
        for col_idx, val in enumerate(row, 1):
            cell = ws_sec.cell(row=curr_row, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        curr_row += 1

    # 2. KẾT QUẢ CHẠY TEST CASE BẢO MẬT
    curr_row += 1
    ws_sec.cell(row=curr_row, column=1, value="2. BẢNG KẾT QUẢ KIỂM THỬ BẢO MẬT THỰC TẾ").font = font_section
    ws_sec.row_dimensions[curr_row].height = 20
    curr_row += 1

    headers_sec_run = ["STT", "Tên Kịch Bản Bảo Mật", "Yêu cầu HTTP / Endpoint", "Dữ liệu payload kiểm thử", "Kết quả mong đợi", "Kết quả thực tế (Newman)", "Trạng thái"]
    ws_sec.row_dimensions[curr_row].height = 28
    for col_num, h in enumerate(headers_sec_run, 1):
        cell = ws_sec.cell(row=curr_row, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Parse collection Security
    sec_testcases = []
    if os.path.exists(security_collection_path):
        try:
            with open(security_collection_path, "r", encoding="utf-8") as f:
                coll = json.load(f)
                idx = 1
                for folder in coll.get("item", []):
                    folder_name = folder.get("name")
                    for req_item in folder.get("item", []):
                        req = req_item.get("request", {})
                        name = req_item.get("name", "")
                        method = req.get("method", "POST")
                        path = "/".join(req.get("url", {}).get("path", []))
                        body_raw = req.get("body", {}).get("raw", "")
                        
                        expected = "API phản hồi thành công và dữ liệu được làm sạch."
                        if "rate limit" in name.lower() or "login liên tiếp" in name.lower():
                            expected = "Các request đầu trả về 401/200, từ request thứ 6 trong 1 phút trả về 429 Too Many Requests."
                        elif "sql injection" in name.lower():
                            expected = "API không bị lỗi DB (không crash 500) và từ chối truy cập hoặc trả về danh sách rỗng."

                        res_info = security_actual.get(name, {"status": "Untested", "log": "Chưa chạy kiểm thử"})

                        sec_testcases.append({
                            "stt": idx,
                            "name": name,
                            "endpoint": f"[{method}] /{path}",
                            "body": body_raw,
                            "expected": expected,
                            "actual": res_info["log"],
                            "status": res_info["status"]
                        })
                        idx += 1
        except Exception as e:
            print(f"⚠️ Lỗi khi trích xuất collection Security: {e}")

    curr_row += 1
    for tc in sec_testcases:
        ws_sec.row_dimensions[curr_row].height = 65
        ws_sec.cell(row=curr_row, column=1, value=tc["stt"])
        ws_sec.cell(row=curr_row, column=2, value=tc["name"])
        ws_sec.cell(row=curr_row, column=3, value=tc["endpoint"])
        ws_sec.cell(row=curr_row, column=4, value=tc["body"])
        ws_sec.cell(row=curr_row, column=5, value=tc["expected"])
        ws_sec.cell(row=curr_row, column=6, value=tc["actual"])
        
        status_cell = ws_sec.cell(row=curr_row, column=7, value=tc["status"])

        # Format styles
        for col_idx in range(1, 8):
            cell = ws_sec.cell(row=curr_row, column=col_idx)
            cell.font = font_body
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in [1, 3, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color status
        if tc["status"] == "Passed":
            status_cell.fill = fill_passed
            status_cell.font = font_passed
        elif tc["status"] == "Failed":
            status_cell.fill = fill_failed
            status_cell.font = font_failed
        else:
            status_cell.fill = fill_untested
            status_cell.font = font_untested

        curr_row += 1

    # Set column widths
    widths_sec = {"A": 8, "B": 24, "C": 22, "D": 32, "E": 20, "F": 35, "G": 12}
    for col, w in widths_sec.items():
        ws_sec.column_dimensions[col].width = w

    # Lưu Workbook Excel
    wb.save(excel_output)
    print(f"🎉 Đã xuất thành công file báo cáo Excel Học thuật tại: {excel_output}")
    return excel_output

def upload_to_gdrive(file_path):
    """Tải file Excel lên Google Drive chung của team"""
    from google.oauth2.credentials import Credentials
    
    folder_id = os.getenv("GDRIVE_FOLDER_ID")
    if not folder_id:
        print("⚠️ Thiếu cấu hình thư mục Google Drive (GDRIVE_FOLDER_ID). Không thể upload!")
        return

    file_name = os.path.basename(file_path)
    file_metadata = {
        'name': file_name,
        'parents': [folder_id]
    }
    
    media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

    refresh_token = os.getenv("GDRIVE_REFRESH_TOKEN")
    client_id = os.getenv("GDRIVE_CLIENT_ID")
    client_secret = os.getenv("GDRIVE_CLIENT_SECRET")

    if refresh_token and client_id and client_secret:
        print("🔄 Đang thử upload bằng phương thức OAuth2...")
        try:
            creds = Credentials(
                token=None,
                refresh_token=refresh_token,
                token_uri="https://oauth2.googleapis.com/token",
                client_id=client_id,
                client_secret=client_secret
            )
            service = build('drive', 'v3', credentials=creds)
            service.about().get(fields="user").execute()
            
            new_file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            print(f"✅ Đã tải thành công file báo cáo lên Google Drive bằng OAuth2!")
            print(f" Tên file: {file_name} | ID: {new_file.get('id')}")
            return
        except Exception as oauth_error:
            print(f"⚠️ Phương thức OAuth2 gặp lỗi: {oauth_error}")
            print("🔄 Đang chuyển sang phương thức dự phòng Service Account...")

    key_path = ".gdrive_key.json"
    raw_key = os.getenv("GDRIVE_SERVICE_ACCOUNT_KEY_RAW")
    if raw_key:
        with open(".gdrive_key.json", "w") as temp_file:
            temp_file.write(raw_key)
        key_path = ".gdrive_key.json"

    if os.path.exists(key_path):
        try:
            print("🔄 Đang tiến hành upload bằng Service Account...")
            scopes = ['https://www.googleapis.com/auth/drive']
            creds = service_account.Credentials.from_service_account_file(key_path, scopes=scopes)
            service = build('drive', 'v3', credentials=creds)
            
            new_file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            print(f"✅ Đã tải thành công file báo cáo lên Google Drive bằng Service Account!")
            print(f" Tên file: {file_name} | ID: {new_file.get('id')}")
            
            if os.getenv("GDRIVE_SERVICE_ACCOUNT_KEY_RAW") and os.path.exists(".gdrive_key.json"):
                try:
                    os.remove(".gdrive_key.json")
                except Exception:
                    pass
            return
        except Exception as sa_error:
            print(f"❌ Lỗi khi upload bằng Service Account: {sa_error}")
    else:
        print("❌ Không tìm thấy thông tin xác thực Google Drive hợp lệ!")

if __name__ == "__main__":
    report_file = generate_academic_report()
    if report_file:
        upload_to_gdrive(report_file)
