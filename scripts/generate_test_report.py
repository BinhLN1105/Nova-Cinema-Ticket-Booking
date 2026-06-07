import os
import sys
import io
import json
import datetime

# Fix encoding cho Windows console (tránh crash khi in emoji Unicode)
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from dotenv import load_dotenv

# Load cấu hình từ file .env ở local (nếu có)
load_dotenv()

# ── CẤU HÌNH & MAPPING THÀNH VIÊN THEO LUỒNG XOAY VÒNG CHÉO ──────────────────
# Mapping: Phân hệ API -> Ai là Tester kiểm thử -> Ai là Dev sửa BUG
TESTER_MAPPING = {
    "auth": {"tester": "Nguyên Vũ", "dev": "Duy Tuấn"},
    "movies": {"tester": "Minh Triết", "dev": "Minh Trí"},
    "cinemas": {"tester": "Minh Thắng", "dev": "Nhật Bình"},
    "bookings": {"tester": "Minh Trí", "dev": "Minh Thắng"},
    "payments": {"tester": "Duy Tuấn", "dev": "Minh Triết"},
    "utilities": {"tester": "Nhật Bình", "dev": "Nguyên Vũ"}
}

# Mapping: Tên folder trong Postman Collection -> Key module tương ứng
FOLDER_TO_MODULE = {
    "auth & security": "auth",
    "movies & showtimes": "movies",
    "cinemas, screens & seat layout": "cinemas",
    "booking & check-in flow": "bookings",
    "payments & wallet flow": "payments",
    "utilities": "utilities"
}

def get_module_key(url_path, folder_name=""):
    """Xác định phân hệ dựa trên tên Folder cha (ưu tiên) hoặc URL đường dẫn API (dự phòng)"""
    # Ưu tiên 1: Phân loại dựa trên tên Folder cha trong Postman Collection
    folder_lower = folder_name.lower()
    if folder_lower in FOLDER_TO_MODULE:
        return FOLDER_TO_MODULE[folder_lower]

    # Ưu tiên 2: Nếu folder không khớp, fallback sang phân loại theo URL
    path_lower = url_path.lower()
    if "auth" in path_lower:
        return "auth"
    elif "movies" in path_lower or "genres" in path_lower:
        return "movies"
    elif "cinemas" in path_lower or "screens" in path_lower:
        return "cinemas"
    elif "bookings" in path_lower or "check-in" in path_lower:
        return "bookings"
    elif "payments" in path_lower:
        return "payments"
    return "utilities"

def generate_report():
    collection_path = "qa-tests/postman/NOVATicket.postman_collection.json"
    results_path = "baocaoLocal/postman-report.json"
    
    # Sinh tên file Excel có kèm dấu vết ngày tháng năm giờ phút chạy test
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    
    # Đảm bảo thư mục baocaoLocal tồn tại trước khi lưu
    output_dir = "baocaoLocal"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    excel_output = os.path.join(output_dir, f"NovaTicket_API_Test_Report_{timestamp}.xlsx")

    if not os.path.exists(collection_path):
        print(f"❌ Không tìm thấy file Postman collection tại {collection_path}")
        return None

    # 1. Đọc kết quả chạy thực tế (nếu có từ Newman run)
    actual_statuses = {}
    if os.path.exists(results_path):
        try:
            with open(results_path, "r", encoding="utf-8") as f:
                report_data = json.load(f)
                for execution in report_data.get("run", {}).get("executions", []):
                    item_name = execution.get("item", {}).get("name")
                    # Mặc định Passed nếu không có assertions nào fail
                    is_failed = False
                    for assertion in execution.get("assertions", []):
                        if assertion.get("error"):
                            is_failed = True
                            break
                    actual_statuses[item_name] = "Failed" if is_failed else "Passed"
            print("📊 Đã nạp thành công kết quả chạy test thực tế từ Newman!")
        except Exception as e:
            print(f"⚠️ Không thể đọc kết quả chạy thực tế: {e}")

    # 2. Đọc file Postman Collection để trích xuất testcases
    with open(collection_path, "r", encoding="utf-8") as f:
        collection = json.load(f)

    testcases = []

    def extract_requests(items, folder_name):
        """Đệ quy duyệt qua tất cả các cấp thư mục con để trích xuất toàn bộ request"""
        for request_item in items:
            # Nếu item này là sub-folder (có chứa "item" con bên trong) → đệ quy xuống sâu hơn
            if "item" in request_item and request_item.get("item") is not None:
                sub_folder_name = request_item.get("name", folder_name)
                extract_requests(request_item["item"], folder_name)
                continue

            # Nếu item này là request thực sự (có chứa "request")
            req_data = request_item.get("request", {})
            if not req_data:
                continue

            req_name = request_item.get("name", "")
            method = req_data.get("method", "GET")
            
            # Trích xuất URL
            url_obj = req_data.get("url", {})
            if isinstance(url_obj, dict):
                url_path = "/".join(url_obj.get("path", []))
            else:
                url_path = str(url_obj)

            # Phân loại module và lấy mapping Tester/Developer (ưu tiên theo Folder cha)
            module_key = get_module_key(url_path, folder_name)
            role_info = TESTER_MAPPING.get(module_key, TESTER_MAPPING["utilities"])

            # Trích xuất mã Test Case tự động từ tên (vd: [201] Đăng ký thành công -> TC-AUTH-201)
            tc_id = f"TC-{module_key.upper()}-{method}"
            if "[" in req_name and "]" in req_name:
                code = req_name.split("]")[0].replace("[", "").strip()
                tc_id = f"TC-{module_key.upper()}-{code}"

            # Các bước thực hiện (Steps)
            steps = f"1. Gửi request [{method}] tới endpoint /{url_path}\n"
            body = req_data.get("body", {})
            if body.get("mode") == "raw" and body.get("raw"):
                steps += f"2. Request Body raw:\n{body.get('raw')}"

            # Kết quả mong muốn (Expected Results)
            expected = "1. API phản hồi thành công.\n"
            event_list = request_item.get("event", [])
            for event in event_list:
                if event.get("listen") == "test":
                    execs = event.get("script", {}).get("exec", [])
                    assertions = [line for line in execs if "pm.test" in line]
                    if assertions:
                        expected += "Ràng buộc kiểm thử:\n" + "\n".join([f"• {a.strip()}" for a in assertions])

            # Trạng thái thực tế
            status = actual_statuses.get(req_name, "Untested") # Mặc định Untested nếu chưa chạy test/chưa bật server

            testcases.append({
                "id": tc_id,
                "name": req_name,
                "steps": steps,
                "expected": expected,
                "module": folder_name,
                "tester": role_info["tester"],
                "dev": role_info["dev"],
                "status": status,
                "date": datetime.date.today().strftime("%Y-%m-%d")
            })

    # Duyệt qua các thư mục phân hệ cấp cao nhất trong Collection
    for folder in collection.get("item", []):
        folder_name = folder.get("name", "Utilities")
        extract_requests(folder.get("item", []), folder_name)

    # 3. Sử dụng openpyxl để vẽ bảng Excel sang xịn mịn
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Test Matrix"

    # Thiết lập lưới hiển thị
    ws.views.sheetView[0].showGridLines = True

    # Định dạng Font & Color chủ đạo (Sơn Navy #0D1B2A và Gold #F5C518)
    font_title = Font(name="Inter", size=16, bold=True, color="0D1B2A")
    font_header = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Inter", size=10)
    font_passed = Font(name="Inter", size=10, bold=True, color="2E7D32")
    font_failed = Font(name="Inter", size=10, bold=True, color="C62828")
    font_untested = Font(name="Inter", size=10, bold=True, color="757575")

    fill_header = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    fill_passed = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_failed = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    fill_untested = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Tiêu đề file Excel
    ws.merge_cells("A1:I1")
    ws["A1"] = f"BẢNG ĐÁNH GIÁ CHẤT LƯỢNG API & PHÂN VAI XOAY VÒNG KIỂM THỬ CHÉO - {datetime.date.today().strftime('%d/%m/%Y')}"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Khai báo Header
    headers = [
        "Mã Test Case", "Tên Kịch Bản API", "Phân Hệ", 
        "Các Bước Thực Hiện (Steps)", "Kết Quả Mong Muốn (Expected)", 
        "Tester (Người Test)", "Developer (Người Sửa Bug)", "Trạng Thái", "Ngày Chạy"
    ]
    
    ws.row_dimensions[3].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Ghi dữ liệu vào bảng
    row_num = 4
    for tc in testcases:
        ws.row_dimensions[row_num].height = 60
        
        c_id = ws.cell(row=row_num, column=1, value=tc["id"])
        c_name = ws.cell(row=row_num, column=2, value=tc["name"])
        c_mod = ws.cell(row=row_num, column=3, value=tc["module"])
        c_steps = ws.cell(row=row_num, column=4, value=tc["steps"])
        c_exp = ws.cell(row=row_num, column=5, value=tc["expected"])
        c_tester = ws.cell(row=row_num, column=6, value=tc["tester"])
        c_dev = ws.cell(row=row_num, column=7, value=tc["dev"])
        c_status = ws.cell(row=row_num, column=8, value=tc["status"])
        c_date = ws.cell(row=row_num, column=9, value=tc["date"])

        # Format các ô dữ liệu thông thường
        for col_index in range(1, 10):
            cell = ws.cell(row=row_num, column=col_index)
            cell.font = font_body
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_index in [1, 3, 6, 7, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Định dạng màu sắc riêng cho Trạng thái
        if tc["status"] == "Passed":
            c_status.fill = fill_passed
            c_status.font = font_passed
        elif tc["status"] == "Failed":
            c_status.fill = fill_failed
            c_status.font = font_failed
        else: # Untested
            c_status.fill = fill_untested
            c_status.font = font_untested

        row_num += 1

    # Tự động căn chỉnh độ rộng cột
    column_widths = {
        "A": 16, "B": 30, "C": 20, "D": 45, "E": 45, "F": 18, "G": 18, "H": 14, "I": 14
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Lưu file Excel local
    wb.save(excel_output)
    print(f"🎉 Đã xuất thành công file báo cáo Excel tại: {excel_output}")
    return excel_output

def upload_to_gdrive(file_path):
    """Tải file Excel lên Google Drive chung của team, hỗ trợ cả OAuth2 cá nhân và Service Account với cơ chế Fallback thực tế"""
    from google.oauth2.credentials import Credentials
    
    folder_id = os.getenv("GDRIVE_FOLDER_ID")
    if not folder_id:
        print("⚠️ Thiếu cấu hình thư mục Google Drive (GDRIVE_FOLDER_ID trong file .env hoặc env variables). Không thể upload!")
        return

    # Lấy tên file thực tế từ file_path
    file_name = os.path.basename(file_path)
    file_metadata = {
        'name': file_name,
        'parents': [folder_id]
    }
    
    media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

    # ── PHƯƠNG ÁN A: DÙNG OAUTH2 NHÂN DANH TÀI KHOẢN CÁ NHÂN (Có Refresh Token) ──
    refresh_token = os.getenv("GDRIVE_REFRESH_TOKEN")
    client_id = os.getenv("GDRIVE_CLIENT_ID")
    client_secret = os.getenv("GDRIVE_CLIENT_SECRET")

    if refresh_token and client_id and client_secret:
        print("🔄 Đang thử xác thực và upload bằng phương thức OAuth2 (Tài khoản cá nhân)...")
        try:
            creds = Credentials(
                token=None,
                refresh_token=refresh_token,
                token_uri="https://oauth2.googleapis.com/token",
                client_id=client_id,
                client_secret=client_secret
            )
            service = build('drive', 'v3', credentials=creds)
            
            # Thử thực hiện một truy vấn API siêu nhẹ để verify token thực tế (check mạng)
            service.about().get(fields="user").execute()
            
            # Nếu verify mạng qua được, tiến hành upload!
            new_file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            print(f"✅ Đã tải thành công file báo cáo lên Google Drive bằng OAuth2 (Tài khoản cá nhân)!")
            print(f" Tên file: {file_name} | ID: {new_file.get('id')}")
            return
        except Exception as oauth_error:
            print(f"⚠️ Phương thức OAuth2 gặp lỗi xác thực: {oauth_error}")
            print("🔄 Đang tự động chuyển hướng (fallback) sang phương thức dự phòng (Service Account)...")

    # ── PHƯƠNG ÁN B: DÙNG SERVICE ACCOUNT (Nếu OAuth2 trống hoặc gặp lỗi xác thực ở trên) ──
    key_path = ".gdrive_key.json"
    raw_key = os.getenv("GDRIVE_SERVICE_ACCOUNT_KEY_RAW")
    if raw_key:
        with open(".gdrive_key.json", "w") as temp_file:
            temp_file.write(raw_key)
        key_path = ".gdrive_key.json"

    if os.path.exists(key_path):
        try:
            print("🔄 Đang tiến hành upload bằng Service Account (Tài khoản dịch vụ)...")
            scopes = ['https://www.googleapis.com/auth/drive']
            creds = service_account.Credentials.from_service_account_file(key_path, scopes=scopes)
            service = build('drive', 'v3', credentials=creds)
            
            new_file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            print(f"✅ Đã tải thành công file báo cáo lên Google Drive bằng Service Account (Tài khoản dịch vụ)!")
            print(f" Tên file: {file_name} | ID: {new_file.get('id')}")
            
            # Xóa file key tạm nếu được sinh ra trong phiên chạy CI/CD
            if os.getenv("GDRIVE_SERVICE_ACCOUNT_KEY_RAW") and os.path.exists(".gdrive_key.json"):
                try:
                    os.remove(".gdrive_key.json")
                except Exception:
                    pass
            return
        except Exception as sa_error:
            print(f"❌ Lỗi khi upload bằng Service Account: {sa_error}")
    else:
        print("❌ Không tìm thấy thông tin xác thực Google Drive hợp lệ (OAuth2 thất bại và không tìm thấy file .gdrive_key.json)!")

if __name__ == "__main__":
    report_file = generate_report()
    if report_file:
        upload_to_gdrive(report_file)
